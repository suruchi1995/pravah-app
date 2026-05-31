"""
Generate the client-facing Excel template from the data contract.
Required sheets are highlighted; optional sheets note their default behaviour.
Each sheet has the correct headers + 2 example rows (from Apex) the client overwrites.
Run: python3 scripts/make_template.py
"""
import os, sys, csv
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from backend.data_contract import CONTRACT, REQUIRED_SHEETS, OPTIONAL_SHEETS, DEFAULTS_DOC

DATA = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "datasets")
OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Pravah_Client_Template.xlsx")

REQ = PatternFill("solid", start_color="1B5E20")   # green = required
OPT = PatternFill("solid", start_color="8D6E63")   # brown = optional
HDR = PatternFill("solid", start_color="2E5C8A")
white = Font(bold=True, color="FFFFFF", name="Calibri", size=11)


def example_rows(sheet, cols, n=2):
    """Pull a couple of real Apex rows as examples (without bookkeeping cols)."""
    path = os.path.join(DATA, sheet + ".csv")
    if not os.path.exists(path):
        return []
    out = []
    with open(path) as f:
        for i, r in enumerate(csv.DictReader(f)):
            if i >= n:
                break
            out.append([r.get(c, "") for c in cols])
    return out


def build():
    wb = Workbook()
    idx = wb.active; idx.title = "README"
    idx["A1"] = "PRAVAH — Client Data Template"
    idx["A1"].font = Font(bold=True, size=16, color="1F3A5F")
    idx["A2"] = "Fill the GREEN (required) sheets with your data. BROWN sheets are optional — leave blank to use Pravah defaults."
    idx["A2"].font = Font(italic=True, size=10, color="555555")
    idx["A4"] = "Sheet"; idx["B4"] = "Required?"; idx["C4"] = "Notes"
    for c in ["A4", "B4", "C4"]:
        idx[c].font = white; idx[c].fill = HDR
    row = 5
    for s in REQUIRED_SHEETS + OPTIONAL_SHEETS:
        req = s in REQUIRED_SHEETS
        idx.cell(row=row, column=1, value=s)
        rc = idx.cell(row=row, column=2, value="REQUIRED" if req else "optional")
        rc.fill = REQ if req else OPT; rc.font = Font(color="FFFFFF", bold=True)
        note = "Your data." if req else DEFAULTS_DOC.get(s, "Optional.")
        idx.cell(row=row, column=3, value=note)
        row += 1
    idx.column_dimensions["A"].width = 24
    idx.column_dimensions["B"].width = 12
    idx.column_dimensions["C"].width = 70

    for s in REQUIRED_SHEETS + OPTIONAL_SHEETS:
        spec = CONTRACT[s]
        ws = wb.create_sheet(s)
        ws.append(spec["columns"])
        for cell in ws[1]:
            cell.font = white
            cell.fill = REQ if s in REQUIRED_SHEETS else OPT
        for ex in example_rows(s, spec["columns"]):
            ws.append(ex)
        for col in ws.columns:
            m = max((len(str(c.value)) if c.value is not None else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max(m + 2, 12), 36)
        ws.freeze_panes = "A2"

    wb.save(OUT)
    return OUT


if __name__ == "__main__":
    path = build()
    print("Template written:", path)
